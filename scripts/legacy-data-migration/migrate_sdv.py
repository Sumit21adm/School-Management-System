#!/usr/bin/env python3
"""
SDV Data Migration Script
Migrates student and fee data from legacy MySQL dump to Excel format for import.

Usage:
    python migrate_sdv.py --discover    # Scan tables and generate discovery report
    python migrate_sdv.py --validate    # Validate all data before export
    python migrate_sdv.py --export      # Generate Excel files (all sessions)
    python migrate_sdv.py --export --session 2024-2025  # Single session
"""

import re
import os
import json
import argparse
from datetime import datetime, date
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

try:
    import pandas as pd
    from openpyxl import Workbook
except ImportError:
    print("Required packages not found. Install with: pip install pandas openpyxl")
    exit(1)

# =============================================================================
# CONFIGURATION
# =============================================================================

# Paths are now configured via CLI arguments
# SQL_FILE, OUTPUT_DIR removed from global scope


# Comprehensive Fee Type Mapping
# Legacy Name (any case/format) -> System Name
FEE_TYPE_SYSTEM_MAP = {
    'tuition_fee': 'Tuition Fee',
    'tuitionfeedet': 'Tuition Fee',
    'tuition': 'Tuition Fee',
    'tuition fee': 'Tuition Fee',
    
    'computer_fee': 'Computer Fee',
    'computerfinearts': 'Computer Fee',
    'computer': 'Computer Fee',
    'computer fine arts': 'Computer Fee',
    
    'transport_fee': 'Transport Fee',
    'conveyance': 'Transport Fee',
    'transport': 'Transport Fee',
    
    'dev_fee': 'Development Fee',
    'development': 'Development Fee',
    
    'exam_fee': 'Exam Fee',
    'exam': 'Exam Fee',
    
    'lib_fee': 'Library Fee',
    'library': 'Library Fee',
    
    'lab_fee': 'Lab Fee',
    'laboratory': 'Lab Fee',
    'lab': 'Lab Fee',
    
    'fine': 'Late Fee',
    'latefine': 'Late Fee',
    'late fine': 'Late Fee',
    
    'adm_fee': 'Admission Fee',
    'admission fee': 'Admission Fee',
    
    'pre_dues': 'Previous Dues',
    'pre dues': 'Previous Dues',
    'dues': 'Previous Dues',
    
    'smartclassgencharge': 'Smart Class',
    'smart class': 'Smart Class',
    'smart_class': 'Smart Class',
    
    'generator': 'Generator Fee',
    'gen': 'Generator Fee',
    
    'activity': 'Activity Fee',
    'activity fee': 'Activity Fee',
    
    'dress_fee': 'Dress Fee',
    'dressdues': 'Dress Fee',
    
    'hostel_fee': 'Hostel Fee',
    'hostel': 'Hostel Fee',
    
    'other': 'Other Fee',
    'others': 'Other Fee',
    'other fee': 'Other Fee'
}

def map_fee_type(legacy_name: str) -> Optional[str]:
    """Strictly map legacy names to system names."""
    if not legacy_name or legacy_name in PLACEHOLDER_VALUES:
        return None
    
    # Normalize for lookup
    clean_name = legacy_name.strip().lower()
    
    # Check map
    return FEE_TYPE_SYSTEM_MAP.get(clean_name, legacy_name) # Fallback to original if not in map, but we should be thorough

def is_valid_fee(fee_name: str, amount: float) -> bool:
    """Check if a fee record is valid for import."""
    if amount <= 0:
        return False
    if not fee_name or fee_name in PLACEHOLDER_VALUES:
        return False
    return True

# Placeholder values to treat as NULL
PLACEHOLDER_VALUES = {'--Select--', 'None', 'N/A', '-', '', 'null', 'NULL', 'NA'}

# =============================================================================
# DATA CLASSES
# =============================================================================

class ValidationResult:
    def __init__(self):
        self.valid_students = []
        self.valid_receipts = []
        self.valid_discounts = []
        self.warnings = []  # Auto-fixable issues
        self.errors = []    # Blocking issues
        self.orphan_receipts = []

    def add_warning(self, category: str, record_id: str, field: str, original: str, fixed: str):
        self.warnings.append({
            'category': category,
            'id': record_id,
            'field': field,
            'original': original,
            'fixed': fixed
        })

    def add_error(self, category: str, record_id: str, message: str):
        self.errors.append({
            'category': category,
            'id': record_id,
            'message': message
        })

# =============================================================================
# SQL PARSER
# =============================================================================

def parse_sql_file(filepath: str) -> Dict[str, List[Dict]]:
    """Parse SQL dump file and extract data from all tables."""
    tables = {}
    
    with open(filepath, 'r', encoding='latin1') as f:
        content = f.read()
    
    # Find all INSERT statements
    insert_pattern = r"insert\s+into\s+`(\w+)`\s*\([^)]+\)\s*values\s*"
    
    current_pos = 0
    while True:
        match = re.search(insert_pattern, content[current_pos:], re.IGNORECASE)
        if not match:
            break
        
        table_name = match.group(1)
        start = current_pos + match.end()
        
        # Find column names
        col_match = re.search(r"insert\s+into\s+`" + table_name + r"`\s*\(([^)]+)\)", 
                              content[current_pos:current_pos + match.end()], re.IGNORECASE)
        if col_match:
            columns = [c.strip().strip('`') for c in col_match.group(1).split(',')]
        else:
            columns = []
        
        # Parse values - find the end of this INSERT statement
        values_end = content.find(';\n', start)
        if values_end == -1:
            values_end = len(content)
        
        values_str = content[start:values_end]
        
        # Parse individual value tuples
        if table_name not in tables:
            tables[table_name] = {'columns': columns, 'rows': []}
        
        # Extract value tuples using regex
        tuple_pattern = r"\(([^)]+)\)"
        for tuple_match in re.finditer(tuple_pattern, values_str):
            values = parse_value_tuple(tuple_match.group(1))
            if values and columns:
                row = dict(zip(columns, values))
                tables[table_name]['rows'].append(row)
        
        current_pos = values_end + 1
    
    return tables

def parse_value_tuple(value_str: str) -> List[str]:
    """Parse a SQL value tuple into a list of values."""
    values = []
    current = ""
    in_string = False
    escape_next = False
    
    for char in value_str:
        if escape_next:
            current += char
            escape_next = False
        elif char == '\\':
            escape_next = True
        elif char == "'" and not in_string:
            in_string = True
        elif char == "'" and in_string:
            in_string = False
        elif char == ',' and not in_string:
            values.append(current.strip().strip("'"))
            current = ""
        else:
            current += char
    
    if current:
        values.append(current.strip().strip("'"))
    
    return values

# =============================================================================
# DATA CLEANING
# =============================================================================

def clean_phone(phone: str) -> Tuple[str, bool]:
    """Clean phone number. Returns (cleaned, was_modified)."""
    if not phone or phone in PLACEHOLDER_VALUES:
        return '0000000000', True
    
    # Extract digits only
    digits = re.sub(r'\D', '', phone)
    
    # Handle common patterns
    if len(digits) > 10:
        digits = digits[-10:]  # Take last 10 digits
        
    if len(digits) < 10:
         return '0000000000', True

    was_modified = digits != phone
    return digits, was_modified

def clean_date(date_str: str, fallback: str = '01-01-2000') -> Tuple[str, bool]:
    """Clean date string. Returns (cleaned, was_modified)."""
    if not date_str or date_str in PLACEHOLDER_VALUES:
        return fallback, True
    
    # Handle invalid MySQL dates
    if date_str.startswith('0000') or date_str == '0000-00-00':
        return fallback, True
    
    # Try to parse and reformat to DD-MM-YYYY
    try:
        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d']:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%d-%m-%Y'), date_str != dt.strftime('%d-%m-%Y')
            except ValueError:
                continue
    except:
        pass
    
    return fallback, True

def clean_gender(gender: str) -> str:
    """Clean gender field."""
    if not gender or gender in PLACEHOLDER_VALUES:
        return 'Other'
    gender = gender.strip().lower()
    if gender in ['male', 'm']:
        return 'Male'
    elif gender in ['female', 'f']:
        return 'Female'
    return 'Other'

def clean_text(text: str) -> str:
    """Clean text field - trim whitespace, handle placeholders."""
    if not text:
        return ''
    text = text.strip()
    if text in PLACEHOLDER_VALUES:
        return ''
    return ' '.join(text.split())  # Normalize whitespace

def clean_aadhar(aadhar: str) -> Optional[str]:
    """Clean Aadhar number - return None for placeholders."""
    if not aadhar or aadhar in PLACEHOLDER_VALUES:
        return None
    digits = re.sub(r'\D', '', aadhar)
    if len(digits) == 12:
        return digits
    return None

def safe_float(value) -> float:
    """Convert value to float, handling empty/invalid values and mixed types."""
    if value is None or value == '' or value in PLACEHOLDER_VALUES:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(',', ''))
        except (ValueError, TypeError):
            return 0.0
    return 0.0

# =============================================================================
# DATA EXTRACTION
# =============================================================================

def extract_students(tables: Dict) -> Dict[str, List[Dict]]:
    """Extract and clean student data, grouped by session."""
    students_by_session = defaultdict(list)
    
    if 'student_details' not in tables:
        return students_by_session
    
    for row in tables['student_details']['rows']:
        session = row.get('year', '')
        if not session or not re.match(r'\d{4}-\d{4}', session):
            continue
        
        # Clean phone - try Mobile_No first, then extract from address
        phone, _ = clean_phone(row.get('Mobile_No', ''))
        if not phone:
            # Try extracting from address fields
            for addr_field in ['pr1', 'pe1']:
                addr = row.get(addr_field, '')
                phone_match = re.search(r'(\d{10})', addr)
                if phone_match:
                    phone = phone_match.group(1)
                    break
        
        # Build address from components
        addr_parts = [clean_text(row.get('pr1', '')), clean_text(row.get('pr2', ''))]
        address = ', '.join([p for p in addr_parts if p and not re.match(r'^\d{10}$', p)])
        

        # Fix Class Name
        student_class = clean_text(row.get('clss', ''))
        if not student_class or student_class in PLACEHOLDER_VALUES or student_class == '-':
            student_class = 'PASS OUT'
        
        # Determine Status
        status = row.get('status', '').lower()
        if student_class == 'PASS OUT' or student_class == 'Pass Out':
            status = 'alumni'
        else:
             status = 'active' if status == 'active' else 'inactive'

        student = {
            'student_id': row.get('student_id', ''),
            'name': clean_text(row.get('Student_Name', '')),
            'father_name': clean_text(row.get('Father_Name', '')),
            'mother_name': clean_text(row.get('Mother_Name', '')),
            'dob': clean_date(row.get('DOB', ''))[0],
            'gender': clean_gender(row.get('Sex', '')),
            'class': student_class,
            'section': clean_text(row.get('sec', '')) or 'A',
            'roll': clean_text(row.get('roll', '')),
            'admission_date': clean_date(row.get('date', ''))[0],
            'phone': phone,
            'whats_app': phone, # Default WhatsApp to mobile
            'email': clean_text(row.get('email', '')),
            'address': address if address else 'Address Not Available',
            'aadhar': clean_aadhar(row.get('uidNo', '')),
            'category': clean_text(row.get('cate', '')) or 'NA',
            'religion': clean_text(row.get('Religion', '')),
            'status': status,
            'session': session,
            'father_occupation': clean_text(row.get('Father_Occupation', '')),
            'mother_occupation': clean_text(row.get('Mother_Occupation', '')),
            'father_aadhar': clean_aadhar(row.get('Father_Aadhar', '')),
            'mother_aadhar': clean_aadhar(row.get('Mother_Aadhar', '')),
        }
        
        if student['name'] and student['student_id']:
            students_by_session[session].append(student)
            
    # Deduplicate roll numbers within each session
    for session, students in students_by_session.items():
        students_by_session[session] = deduplicate_roll_numbers(students)
    
    return students_by_session

def deduplicate_roll_numbers(students: List[Dict]) -> List[Dict]:
    """Ensure uniqueness of Roll Numbers within Class & Section."""
    # Key: (class, section) -> set of seen rolls
    seen_rolls = defaultdict(set)
    # Key: (class, section) -> list of students with duplicate/invalid rolls
    to_fix = defaultdict(list)
    
    cleaned_students = []
    
    # First pass: Collect valid unique rolls
    for s in students:
        cls = s['class']
        sec = s['section']
        roll = s['roll']
        
        if not roll or roll in PLACEHOLDER_VALUES:
            to_fix[(cls, sec)].append(s)
            continue
            
        key = (cls, sec)
        if roll in seen_rolls[key]:
            to_fix[key].append(s)
        else:
            seen_rolls[key].add(roll)
            cleaned_students.append(s)
            
    # Second pass: Fix duplicates and missing
    for key, bad_students in to_fix.items():
        cls, sec = key
        existing = seen_rolls[key]
        
        # Strategies to resolve duplicates:
        # 1. Prefix with '000' (User request) to mark them explicitly
        # 2. If '000' prefix conflicts or exceeds length, fallback to next available integer
        
        # Determine next available numeric roll for fallback cases
        max_roll = 0
        for r in existing:
            if r.isdigit():
                max_roll = max(max_roll, int(r))
        next_roll = max_roll + 1
        
        for s in bad_students:
            original_roll = s['roll']
            
            # Strategy 1: Try adding '000' prefix
            # Note: Max DB length is 10 chars
            if original_roll and original_roll not in PLACEHOLDER_VALUES:
                candidate = "000" + original_roll
                # Make sure it's unique and fits in DB
                if candidate not in existing and len(candidate) <= 10:
                    s['roll'] = candidate
                    existing.add(candidate)
                    continue
                    
                # If 000 prefix fails (taken or too long), try 0000
                candidate = "0000" + original_roll
                if candidate not in existing and len(candidate) <= 10:
                    s['roll'] = candidate
                    existing.add(candidate)
                    continue
            
            # Strategy 2: Fallback to next available numeric
            while str(next_roll) in existing:
                next_roll += 1
            
            new_roll = str(next_roll)
            s['roll'] = new_roll
            existing.add(new_roll)
            cleaned_students.append(s)
            
    return cleaned_students

def extract_demand_bills(tables: Dict, students_by_session: Dict) -> Dict[str, List[Dict]]:
    """Extract demand bills joining demandbillnew (amounts) and demandbillsec (meta)."""
    bills_by_session = defaultdict(list)
    
    if 'demandbillnew' not in tables or 'demandbillsec' not in tables:
        return bills_by_session

    # Build metadata lookup from demandbillsec
    bill_meta = {}
    for row in tables['demandbillsec']['rows']:
        bill_no = row.get('billNo', '')
        if bill_no:
            bill_meta[bill_no] = {
                'year': row.get('billYear', ''),
                'month': row.get('billmonth', ''),
                'date': clean_date(row.get('currentDate', ''))[0]
            }

    # Build student ID lookup
    all_student_ids = set()
    for session_students in students_by_session.values():
        for s in session_students:
            all_student_ids.add(str(s['student_id']))

    for row in tables['demandbillnew']['rows']:
        bill_no = row.get('BillNo', '')
        student_id = str(row.get('StudentID', ''))
        
        if not bill_no or not student_id or student_id not in all_student_ids:
            continue

        # Get metadata
        meta = bill_meta.get(bill_no, {})
        session = meta.get('year', '')
        bill_date = meta.get('date', '')
        
        # Fallback if session missing in meta (check row itself just in case)
        if not session:
             session = row.get('Year', '') # Unlikely based on schema but safe
        
        if not session:
            # Fallback to current if date matches? Or skip?
            # Let's skip orphan bills without session
            if '202' in bill_date: # minimal guess
                 y = bill_date.split('-')[-1]
                 if y == '2024': session = '2024-2025'
                 elif y == '2025': session = '2025-2026'
            
        if not session:
            continue

        # Extract fee components
        previous_dues = safe_float(row.get('Dues', '0'))
        
        # Map known columns to Fee Types
        # Extract fee components
        # Based on demandbillnew schema
        legacy_cols = [
            'TuitionFee', 'ComputerFineArts', 'TransportFee', 'Conveyance', 'SmartClassGenCharge', 
            'Development', 'Laboratory', 'Library', 'LateFine', 'Others', 
            'Activity', 'Exam', 'DressDues', 'HostelFee'
        ]
        
        for col in legacy_cols:
            amount = safe_float(row.get(col, '0'))
            fee_type = map_fee_type(col)
            
            if is_valid_fee(fee_type, amount):
                bills_by_session[session].append({
                    'student_id': student_id,
                    'bill_no': bill_no,
                    'bill_date': bill_date,
                    'fee_type': fee_type,
                    'amount': amount,
                    'net_amount': amount # Default net
                })
        
        # NOTE: We ignore 'Dues' column from the bill because it represents 
        # cumulative arrears which the system will calculate automatically 
        # from the imported historical bills and receipts. 
        # Adding it here would duplicate the debt every month.
            
    return bills_by_session

def extract_admission_payments(tables: Dict, students_by_session: Dict) -> Dict[str, List[Dict]]:
    """Extract admissionpayment data as fee receipts."""
    receipts_by_session = defaultdict(list)
    
    if 'admissionpayment' not in tables or 'financialmaster' not in tables:
        return receipts_by_session

    # Build session lookup from financialmaster
    # Schema: financialid, financialyear
    year_map = {}
    for row in tables['financialmaster']['rows']:
        fid = row.get('financialid') or row.get('id')
        fyear = row.get('financialyear') or row.get('year')
        if fid and fyear:
            year_map[str(fid)] = fyear

    # Build student ID lookup
    all_student_ids = set()
    for session_students in students_by_session.values():
        for s in session_students:
            all_student_ids.add(str(s['student_id']))

    # Group by transactionId to form receipts
    # receipt_id -> { date, student_id, items: [] }
    # Since we lack a date table, we default to 1st April of the session start year
    
    transactions = defaultdict(lambda: {'items': [], 'student_id': '', 'session': '', 'amount': 0.0})

    for row in tables['admissionpayment']['rows']:
        # Columns: id, transactionId, studentId, description, amount, yearId
        tid = row.get('transactionId')
        sid = str(row.get('studentId'))
        yid = str(row.get('yearId'))
        desc = row.get('description', 'Fee')
        amt = safe_float(row.get('amount', '0'))
        
        if not tid or not sid or sid not in all_student_ids:
            continue
            
        session = year_map.get(yid, '')
        if not session:
            # Try to guess or skip?
            # If yearId is 3, maybe it's 2018-2019 etc.
            # Let's verify with map later.
            continue
            
        key = f"{tid}_{sid}" # composite key just in case
        
        transactions[key]['student_id'] = sid
        transactions[key]['session'] = session
        transactions[key]['items'].append({
            'fee_type': desc,
            'amount': amt
        })
        transactions[key]['amount'] += amt
        transactions[key]['tid'] = tid

    # Convert to standard receipt format
    for key, data in transactions.items():
        session = data['session']
        
        # Generate default date
        # 2018-2019 -> 01-04-2018
        start_year = session.split('-')[0]
        default_date = f"{start_year}-04-01"
        
        # We need to map description to standard fee types if possible
        # Or keep original description?
        # The user wants "Fee Requirements" matched.
        # Let's attempt mapping or pass through.
        # Mappings: 'Tuition Fee' -> 'Tuition Fee', 'Conveyance' -> 'Transport Fee'
        
        fee_map = {
            'Tuition Fee': 'Tuition Fee',
            'Conveyance': 'Transport Fee',
            'Computer Fine Arts': 'Computer Fee',
            'Smart Class': 'Smart Class',
            'Development': 'Development Fee',
            'Laboratory': 'Lab Fee',
            'Library': 'Library Fee',
            'Activity': 'Activity Fee',
            'Generator': 'Generator Fee', # or similar
            'Others': 'Other Fee'
        }
        
        for item in data['items']:
            mapped_type = map_fee_type(item['fee_type'])
            
            if is_valid_fee(mapped_type, item['amount']):
                receipts_by_session[session].append({
                    'student_id': data['student_id'],
                    'receipt_no': f"ADM-{data['tid']}", # Prefix to distinguish
                    'receipt_date': default_date,
                    'fee_type': mapped_type,
                    'amount': item['amount'],
                    'discount': 0, # Admissionpayment usually net
                    'payment_mode': 'CASH', # Default
                    'payment_ref': ''
                })
            
    return receipts_by_session

def extract_modern_transactions(tables: Dict, students_by_session: Dict) -> Dict[str, List[Dict]]:
    """Extract receipts from feetransaction_new (detailed) and feetransaction_newtwo (consolidated)."""
    receipts_by_session = defaultdict(list)
    
    # helper to build ID lookup
    all_student_ids = set()
    for session_students in students_by_session.values():
        for s in session_students:
            all_student_ids.add(str(s['student_id']))

    # 1. feetransaction_new (Has breakdown)
    if 'feetransaction_new' in tables:
        # Schema: id, transaction_id, student_id, receipt_no, year, date, tuition, computer...
        
        # Map columns to fee types
        col_map = {
            'tuition': 'Tuition Fee',
            'computer': 'Computer Fee',
            'smart_class': 'Smart Class',
            'development': 'Development Fee',
            'lab': 'Lab Fee',
            'library': 'Library Fee',
            'latefine': 'Late Fee',
            'others': 'Other Fee',
            'gen': 'Generator Fee',
            'activity': 'Activity Fee',
            'exam': 'Exam Fee',
            'hostel': 'Hostel Fee',
            'conveyance': 'Transport Fee'
        }
        
        for row in tables['feetransaction_new']['rows']:
            session = row.get('year', '')
            sid = str(row.get('student_id', ''))
            
            if not session or not sid or sid not in all_student_ids:
                continue
                
            r_no = row.get('receipt_no', '')
            r_date = clean_date(row.get('date', ''))[0]
            
            # Extract items
            for col, ftype in col_map.items():
                amt = safe_float(row.get(col, '0'))
                if amt > 0:
                    receipts_by_session[session].append({
                        'student_id': sid,
                        'receipt_no': f"REC-{r_no}",
                        'receipt_date': r_date,
                        'fee_type': ftype,
                        'amount': amt,
                        'discount': 0,
                        'payment_mode': 'Cash', # Assumption
                        'payment_ref': ''
                    })

    # 2. feetransaction_newtwo (Consolidated?)
    if 'feetransaction_newtwo' in tables:
        # Schema: transactionId, billNo, datep, financialYear, studentId, totalAmt, paidAmt...
        for row in tables['feetransaction_newtwo']['rows']:
            session = row.get('financialYear', '')
            sid = str(row.get('studentId', ''))
            
            if not session or not sid or sid not in all_student_ids:
                continue
                
            r_no = row.get('billNo', row.get('transactionId', ''))
            r_date = clean_date(row.get('datep', ''))[0]
            paid_amt = safe_float(row.get('paidAmt', '0'))
            
            # If we don't have breakdown columns, we treat as consolidated
            # Note: The table might have columns we didn't see in the CREATE snippet if they were truncated?
            # But relying on what we saw: valid logic is to take paidAmt.
            
            if paid_amt > 0:
                receipts_by_session[session].append({
                    'student_id': sid,
                    'receipt_no': f"REC2-{r_no}",
                    'receipt_date': r_date,
                    'fee_type': 'Tuition Fee', # Defaulting to Tuition as safe bet, or 'Consolidated Fee'
                    # User requested specific fee types matching legacy names. 
                    # 'Consolidated Fee' might not exist in their system.
                    # 'Tuition Fee' is safest for "general payment".
                    'amount': paid_amt,
                    'discount': 0,
                    'payment_mode': row.get('paymode', 'Cash'),
                    'payment_ref': row.get('chequeNo', '')
                })

    return receipts_by_session

def extract_fee_receipts(tables: Dict, students_by_session: Dict) -> Dict[str, List[Dict]]:
    """Extract fee receipts, grouped by session."""
    receipts_by_session = defaultdict(list)
    
    if 'feereceipt' not in tables:
        return receipts_by_session
    
    # Build student ID lookup
    all_student_ids = set()
    for session_students in students_by_session.values():
        for s in session_students:
            all_student_ids.add(str(s['student_id']))
    
    for row in tables['feereceipt']['rows']:
        session = row.get('year', '')
        student_id = str(row.get('student_id', ''))
        
        if not session or not student_id:
            continue
        
        # Check if student exists
        if student_id not in all_student_ids:
            continue  # Orphan receipt - skip
        
        receipt_no = row.get('feereceipt_no', row.get('feereceipt', ''))
        receipt_date = clean_date(row.get('rdate', ''))[0]
        payment_mode = row.get('paymode', 'Cash')
        if payment_mode in PLACEHOLDER_VALUES:
            payment_mode = 'Cash'
        
        # Extract individual fee amounts
        for legacy_col in ['adm_fee', 'tuition_fee', 'computer_fee', 'transport_fee', 'dev_fee', 'exam_fee', 'lib_fee', 'lab_fee', 'fine', 'other', 'pre_dues']:
            amount = safe_float(row.get(legacy_col, '0'))
            fee_type = map_fee_type(legacy_col)
            
            if is_valid_fee(fee_type, amount):
                receipts_by_session[session].append({
                    'student_id': student_id,
                    'receipt_no': receipt_no,
                    'receipt_date': receipt_date,
                    'fee_type': fee_type,
                    'amount': amount,
                    'discount': 0,
                    'payment_mode': payment_mode,
                    'payment_ref': row.get('check_ddNo', ''),
                })
    
    return receipts_by_session

def extract_discounts(tables: Dict, students_by_session: Dict) -> Dict[str, List[Dict]]:
    """Extract student discounts from concessiontable."""
    discounts_by_session = defaultdict(list)
    
    if 'concessiontable' not in tables:
        return discounts_by_session
    
    for row in tables['concessiontable']['rows']:
        session = row.get('Year', row.get('Fin_Year', ''))
        student_id = str(row.get('StudentID', ''))
        
        if not session or not student_id:
            continue
        
        # Extract discount amounts for each fee type
        for legacy_col in ['TuitionFee', 'ComputerFineArts', 'SmartClass', 'Development', 'Laboratory', 'Library', 'LateFine', 'Others', 'Generator', 'Activity', 'Exam']:
            amount = safe_float(row.get(legacy_col, '0'))
            fee_type = map_fee_type(legacy_col)
            
            if is_valid_fee(fee_type, amount):
                discounts_by_session[session].append({
                    'student_id': student_id,
                    'fee_type': fee_type,
                    'discount_amount': amount,
                    'discount_type': 'Fixed',
                    'reason': 'Migrated from legacy system',
                })
    
    return discounts_by_session

# =============================================================================
# VALIDATION
# =============================================================================

def validate_data(students: Dict, receipts: Dict, discounts: Dict) -> ValidationResult:
    """Validate all extracted data."""
    result = ValidationResult()
    
    # Build student ID set per session
    student_ids_by_session = {}
    for session, session_students in students.items():
        student_ids_by_session[session] = {str(s['student_id']) for s in session_students}
        result.valid_students.extend(session_students)
    
    all_student_ids = set()
    for ids in student_ids_by_session.values():
        all_student_ids.update(ids)
    
    # Validate receipts
    for session, session_receipts in receipts.items():
        for r in session_receipts:
            if str(r['student_id']) not in all_student_ids:
                result.add_error('receipt', r['receipt_no'], 
                               f"Student {r['student_id']} not found")
                result.orphan_receipts.append(r)
            else:
                result.valid_receipts.append(r)
    
    # Validate discounts
    for session, session_discounts in discounts.items():
        for d in session_discounts:
            if str(d['student_id']) not in all_student_ids:
                result.add_error('discount', d['student_id'], 
                               f"Student {d['student_id']} not found")
            else:
                result.valid_discounts.append(d)
    
    return result

# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_excel(session: str, students: List[Dict], receipts: List[Dict], 
                   bills: List[Dict], discounts: List[Dict], output_dir: str, **kwargs) -> str:
    """Generate Excel file by copying template and populating data."""
    os.makedirs(output_dir, exist_ok=True)
    safe_session = session.replace('/', '-').replace(' ', '_')
    filename = f"Migration_{safe_session}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Check if template exists
    template_path = "SDV Data Migration/data_migration_template.xlsx"
    if not os.path.exists(template_path):
        print(f"⚠️ Template not found at {template_path}, creating basic Workbook...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
    else:
        from shutil import copyfile
        copyfile(template_path, filepath)
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
    
    # Helper to append data to worksheet based on headers
    def populate_sheet(ws_name, data, col_mapping):
        if not data or ws_name not in wb.sheetnames:
            return
            
        ws = wb[ws_name]
        
        # Get headers from row 1
        headers = {cell.value: i+1 for i, cell in enumerate(ws[1]) if cell.value}
        
        # Start appending from row 2 onwards (overwriting sample if any)
        current_row = 2
        
        for record in data:
            for key, val in record.items():
                header_name = col_mapping.get(key)
                if header_name and header_name in headers:
                    col_idx = headers[header_name]
                    cell = ws.cell(row=current_row, column=col_idx)
                    
                    # Format dates as DD-MM-YYYY
                    if isinstance(val, (date, datetime)):
                         cell.value = val.strftime('%d-%m-%Y')
                    elif isinstance(val, str) and len(str(val)) == 10 and '-' in str(val): # YYYY-MM-DD
                        try:
                            if val[4] == '-' and val[7] == '-':
                                d_obj = datetime.strptime(val, '%Y-%m-%d')
                                cell.value = d_obj.strftime('%d-%m-%Y')
                            else:
                                cell.value = val
                        except:
                            cell.value = val
                    else:
                        cell.value = val
            current_row += 1
            
    # Mappings
    student_map = {
        'student_id': 'Student ID *',
        'name': 'Name *',
        'father_name': 'Father Name *',
        'mother_name': 'Mother Name *',
        'dob': 'DOB (DD-MM-YYYY) *',
        'gender': 'Gender *',
        'class': 'Class *',
        'section': 'Section *',
        'roll': 'Roll Number',
        'admission_date': 'Admission Date (DD-MM-YYYY) *',
        'phone': 'Phone *',
        'email': 'Email',
        'address': 'Address *',
        'aadhar': 'Student Aadhar',
        'category': 'Category',
        'religion': 'Religion',
        'status': 'Status',
        'session': 'Session Name',
        'apaar_id': 'APAAR ID',
        'father_occupation': 'Father Occ.',
        'father_aadhar': 'Father Aadhar',
        'father_pan': 'Father PAN',
        'mother_occupation': 'Mother Occ.',
        'mother_aadhar': 'Mother Aadhar',
        'mother_pan': 'Mother PAN',
        'guardian_rel': 'Guardian Rel',
        'guardian_name': 'Guardian Name',
        'guardian_phone': 'Guardian Phone',
        'guardian_email': 'Guardian Email',
        'guardian_aadhar': 'Guardian Aadhar',
        'guardian_pan': 'Guardian PAN',
        'guardian_address': 'Guardian Address',
        'whats_app': 'WhatsApp No'
    }
    
    receipt_map = {
        'student_id': 'Student ID *',
        'receipt_no': 'Receipt No *',
        'receipt_date': 'Receipt Date (DD-MM-YYYY) *',
        'fee_type': 'Fee Type *',
        'amount': 'Amount *',
        'discount': 'Discount',
        'net_amount': 'Net Amount *',
        'payment_mode': 'Payment Mode *',
        'payment_ref': 'Payment Ref',
        'collected_by': 'Collected By',
        'remarks': 'Remarks',
        'bill_no': 'Bill No (if against bill)'
    }
    
    bill_map = {
        'student_id': 'Student ID *',
        'bill_no': 'Bill No *',
        'bill_date': 'Bill Date (DD-MM-YYYY) *',
        'due_date': 'Due Date (DD-MM-YYYY) *',
        'month': 'Month (1-12) *',
        'year': 'Year *',
        'fee_type': 'Fee Type *',
        'amount': 'Amount *',
        'discount': 'Discount',
        'previous_dues': 'Previous Dues',
        'late_fee': 'Late Fee',
        'net_amount': 'Net Amount *',
        'paid_amount': 'Paid Amount',
        'status': 'Status *'
    }
    
    discount_map = {
        'student_id': 'Student ID *',
        'fee_type': 'Fee Type *',
        'discount_type': 'Discount Type *',
        'discount_amount': 'Discount Value *',
        'reason': 'Reason',
        'approved_by': 'Approved By',
        'session': 'Session Name'
    }
    
    history_map = {
        'student_id': 'Student ID *',
        'session': 'Session *',
        'class': 'Class *',
        'section': 'Section *',
        'roll': 'Roll Number',
        'status': 'Status *',
        'final_result': 'Final Result'
    }

    # Populate Sheets
    if students: populate_sheet('Students', students, student_map)
    
    if receipts:
        for r in receipts:
            if 'net_amount' not in r:
                r['net_amount'] = safe_float(r.get('amount', 0)) - safe_float(r.get('discount', 0))
            if 'collected_by' not in r: r['collected_by'] = 'Migration'
            if 'remarks' not in r: r['remarks'] = 'Legacy Import'
        populate_sheet('Fee_Receipts', receipts, receipt_map)

    if bills:
        for b in bills:
            if 'net_amount' not in b: b['net_amount'] = b.get('amount', 0)
            if 'status' not in b: b['status'] = 'PENDING'
            if 'due_date' not in b: b['due_date'] = b.get('bill_date')
            # Extract month/year from bill_date
            date_str = str(b.get('bill_date', ''))
            try:
                if '-' in date_str:
                    parts = date_str.split('-')
                    if len(parts[0]) == 4: # YYYY-MM-DD
                        b['year'] = int(parts[0])
                        b['month'] = int(parts[1])
                    else: # DD-MM-YYYY
                        b['year'] = int(parts[2])
                        b['month'] = int(parts[1])
            except:
                b['month'] = 4
                b['year'] = 2024
        populate_sheet('Demand_Bills', bills, bill_map)
        
    if discounts: 
        for d in discounts:
             if 'approved_by' not in d: d['approved_by'] = 'Administrator'
             if 'discount_type' not in d: d['discount_type'] = 'Fixed'
        populate_sheet('Discounts', discounts, discount_map)
    
    if kwargs.get('history'): 
        populate_sheet('Academic_History', kwargs['history'], history_map)

    wb.save(filepath)
    return filepath


# Consolidated Excel generation logic.


def generate_consolidated_excel(all_data: Dict[str, List[Dict]], output_dir: str) -> str:
    """Generate a single consolidated Excel file for all sessions."""
    # Flatten data
    all_students = []
    all_receipts = []
    all_bills = []
    all_discounts = []
    
    for session_data in all_data['students'].values(): all_students.extend(session_data)
    for session_data in all_data['receipts'].values(): all_receipts.extend(session_data)
    for session_data in all_data['bills'].values(): all_bills.extend(session_data)
    for session_data in all_data['discounts'].values(): all_discounts.extend(session_data)
    
    return generate_excel("Consolidated", all_students, all_receipts, all_bills, all_discounts, output_dir)

# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='SDV Data Migration Script')
    parser.add_argument('--input', '-i', default='SdvData17Jan2026.sql', help='Input SQL dump file path')
    parser.add_argument('--output', '-o', default='output', help='Output directory for Excel files')
    parser.add_argument('--discover', action='store_true', help='Scan tables and generate discovery report')
    parser.add_argument('--validate', action='store_true', help='Validate all data before export')
    parser.add_argument('--export', action='store_true', help='Generate Excel files')
    parser.add_argument('--session', help='Export specific session (e.g., "2024-2025")')
    
    args = parser.parse_args()
    
    # 1. Setup Paths
    if not os.path.exists(args.input):
        print(f"Error: Input file '{args.input}' not found.")
        print("Please specify the path to the SQL dump file using --input")
        exit(1)
        
    os.makedirs(args.output, exist_ok=True)
    
    print(f"Loading data from {args.input}...")
    tables = parse_sql_file(args.input)
    print(f"Parsed {len(tables)} tables.")
    
    # 2. Extract Data
    print("Extracting students...")
    students = extract_students(tables)
    total_students = sum(len(s) for s in students.values())
    print(f"Found {total_students} students across {len(students)} sessions.")
    
    print("Extracting receipts...")
    modern_receipts = extract_modern_transactions(tables, students)
    legacy_receipts = extract_fee_receipts(tables, students)
    admission_receipts = extract_admission_payments(tables, students)
    
    # Combine receipts
    receipts = defaultdict(list)
    for s, data in modern_receipts.items(): receipts[s].extend(data)
    for s, data in legacy_receipts.items(): receipts[s].extend(data)
    for s, data in admission_receipts.items(): receipts[s].extend(data)
    
    total_receipts = sum(len(r) for r in receipts.values())
    print(f"Found {total_receipts} fee receipts.")
    
    print("Extracting demand bills...")
    bills = extract_demand_bills(tables, students)
    total_bills = sum(len(b) for b in bills.values())
    print(f"Found {total_bills} demand bills.")
    
    print("Extracting discounts...")
    discounts = extract_discounts(tables, students)
    total_discounts = sum(len(d) for d in discounts.values())
    print(f"Found {total_discounts} discount records.")
    
    # 3. Discovery Report
    if args.discover:
        report_path = os.path.join(args.output, "discovery_report.txt")
        with open(report_path, 'w') as f:
            f.write("SDV Data Migration - Discovery Report\n")
            f.write("=====================================\n\n")
            
            f.write(f"Source File: {args.input}\n")
            f.write(f"Generated: {datetime.now()}\n\n")
            
            f.write("1. Data Summary\n")
            f.write(f"   Total Students: {total_students}\n")
            f.write(f"   Total Receipts: {total_receipts}\n")
            f.write(f"   Total Demand Bills: {total_bills}\n")
            f.write(f"   Total Discounts: {total_discounts}\n\n")
            
            f.write("2. Sessions Found:\n")
            for session in sorted(students.keys()):
                 count = len(students[session])
                 f.write(f"   - {session}: {count} students\n")
                 
        print(f"Discovery report generated at {report_path}")
        
    # 4. Validation
    if args.validate or args.export:
        print("Validating data...")
        validation_result = validate_data(students, receipts, discounts)
        
        print(f"Validation complete: {len(validation_result.errors)} errors, {len(validation_result.warnings)} warnings.")
        
        # Save validation log
        log_path = os.path.join(args.output, "validation_log.json")
        with open(log_path, 'w') as f:
             json.dump({
                 'errors': validation_result.errors,
                 'warnings': validation_result.warnings,
                 'orphan_receipts': validation_result.orphan_receipts
             }, f, indent=2)
             
        if validation_result.errors:
            print(f"Errors found! Check {log_path} for details.")
    
    # 5. Export
    if args.export:
        print("Generating Excel files...")
        
        sessions_to_export = [args.session] if args.session else students.keys()
        
        output_stats = []
        
        for session in sessions_to_export:
            if session not in students:
                print(f"Warning: Session {session} not found in data.")
                continue
            
            print(f"\nProcessing Session: {session}")
            print(f"     Total Students: {len(students[session])}")
            print(f"     Total Receipts: {len(receipts[session])}")
            print(f"     Total Bills: {len(bills[session])}")
            
            filepath = generate_excel(session, students[session], receipts[session], 
                                    bills[session], discounts[session], args.output)
            print(f"  ✅ Saved: {filepath}")
        
        # Generate Consolidated File
        if not args.session:
            print("\n📚 Generating Consolidated Migration File (All Sessions)...")
            cons_path = generate_consolidated_excel({
                'students': students, 'receipts': receipts, 'bills': bills, 'discounts': discounts
            }, args.output)
            print(f"  ✅ Saved Consolidated: {cons_path}")
        
        print("\n🎉 Export complete!")
        return
    
    # Default: show help if no action is specified
    if not any([args.discover, args.validate, args.export]):
        parser.print_help()

if __name__ == '__main__':
    main()
